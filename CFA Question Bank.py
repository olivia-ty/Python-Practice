import random
import time
import openpyxl

# 读取 Workbook, Sheet, Cell
Qbank = openpyxl.load_workbook("CFA Questions Bank.xlsx", read_only=True)
sheet = Qbank['Sheet1']

# 询问是否开始测试
request1 = format(input("\nAre you ready to start the quiz now?\nYes/No\n")).upper()

if request1 == 'YES':

    print("\nGreat, let's start!")
    time.sleep(1)

    # 出题循环
    list = list(range(2, sheet.max_row))

    while len(list) > 0:

        # 设置生成随机数范围
        questionNo = random.choice(list)
        question = sheet.cell(questionNo, 3)
        solution = sheet.cell(questionNo, 4)

        # 打印题目
        print("\nQ" + str(questionNo) + ". " + question.value)
        time.sleep(2)

        # 打印答案
        request2 = format(input("\nSolution(S)/Pass(P)\n")).upper()

        if request2 == 'S':
            print("\n" + solution.value)
            time.sleep(2)

            request3 = format(input("\nInput 'Yes' when you are ready for the next question\n")).upper()

        list.remove(questionNo)

    if len(list) == 0:
        print("Congratulations! Quiz completed")

elif request1 == 'NO':
    print("It's okay, maybe next time!")